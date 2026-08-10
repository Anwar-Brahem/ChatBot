# -*- coding: utf-8 -*-
"""
excel_generator.py - Genere le fichier FOR-054-Multi-J-SOS Analysis rempli
a partir du rapport JSON de l'analyse video.
"""

import json
import shutil
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, Color
from openpyxl.utils import get_column_letter
import openpyxl.cell.cell
import re

# Constants from config (copied here to avoid circular import)
TEMPLATE_NAME = "FOR-054-Multi-J-SOS Analysis.xlsx"
TEMPLATE_SHEET = "Template"

# Column mapping for operations table
COL_STEP_NUM_START = 1       # A
COL_CP_CS_START = 4          # D
COL_DESCRIPTION_START = 6    # F
COL_PHOTO_START = 15         # O
COL_PHOTO_END = 23           # W (merged O:W)
COL_MAIN_STEP_START = 24     # X
COL_CYCLE_TIME = 33          # AG (single column)
COL_KEY_POINTS_START = 34    # AH
COL_REASON_START = 43        # AQ

# Row constants
FIRST_DATA_ROW = 40
MAX_ROWS_PER_PAGE = 22  # Rows 40-61 = 22 operations
HEADER_ROW = 39


def copy_template_to_output(output_path: Path):
    """Copie le template vers le dossier de sortie."""
    template_path = Path(TEMPLATE_NAME)
    if not template_path.exists():
        # Try to find it in parent directories
        for parent in [Path("."), Path(".."), Path("../..")]:
            candidate = parent / TEMPLATE_NAME
            if candidate.exists():
                template_path = candidate
                break

    if not template_path.exists():
        raise FileNotFoundError(f"Template non trouve : {TEMPLATE_NAME}")

    shutil.copy2(template_path, output_path)
    return output_path


def insert_image_into_cell(ws, image_path: str, row: int, start_col: int, row_height: float = 80):
    """Insere une image dans une cellule fusionnee O:W sans deborder."""
    try:
        img = XLImage(image_path)

        max_width = 480
        max_height = 90

        orig_width, orig_height = img.width, img.height
        ratio = min(max_width / orig_width, max_height / orig_height, 1.0)

        img.width = int(orig_width * ratio)
        img.height = int(orig_height * ratio)

        anchor_cell = f"{get_column_letter(start_col)}{row}"
        img.anchor = anchor_cell

        ws.add_image(img)
        ws.row_dimensions[row].height = max(row_height, img.height + 10)
        return True
    except Exception as e:
        print(f"  [WARN] Impossible d'inserer l'image {image_path}: {e}")
        return False


def calculate_row_height(text_length, base_height=15, chars_per_line=50, min_height=60, max_height=200):
    """Calcule la hauteur de ligne en fonction de la longueur du texte (int)."""
    if not text_length or text_length == 0:
        return min_height
    estimated_lines = max(1, text_length // chars_per_line)
    height = base_height * estimated_lines + 20
    return min(max(height, min_height), max_height)


def fill_operation_row(ws, row: int, step_num: int, data: dict, image_path: str = None):
    """Remplit une ligne d'operation dans le template avec hauteur auto."""

    # 1. N° Etape Principale (A:C)
    step_cell = ws.cell(row=row, column=COL_STEP_NUM_START, value=step_num)
    step_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    step_cell.font = Font(name='Calibri', size=11)

    # 2. CP/CS (D:E)
    cp_cs = data.get("cp_cs", "Non")
    cp_cs_cell = ws.cell(row=row, column=COL_CP_CS_START, value=cp_cs)
    cp_cs_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cp_cs_cell.font = Font(name='Calibri', size=11)

    # 3. Description du mode opératoire (F:N)
    description = data.get("description_complete", "")
    key_points = data.get("points_cles", "")
    desc_cell = ws.cell(row=row, column=COL_DESCRIPTION_START)
    apply_keyword_formatting(desc_cell, description, key_points)
    desc_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

    # 4. Complément du mode opératoire - Photo (O:W)
    if image_path and Path(image_path).exists():
        insert_image_into_cell(ws, image_path, row, COL_PHOTO_START, row_height=120)
    else:
        photo_cell = ws.cell(row=row, column=COL_PHOTO_START, value="[Photo non disponible]")
        photo_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # 5. Étapes principales (X:AF) - FORMAT TITRE
    main_step_raw = data.get("etape_principale_resume", "")
    main_step = convert_to_title_format(main_step_raw)
    main_cell = ws.cell(row=row, column=COL_MAIN_STEP_START, value=main_step)
    main_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    main_cell.font = Font(name='Calibri', size=11, bold=True)

    # 6. T/C (AG)
    cycle_time = data.get("temps_cycle_estime", "3s")
    tc_cell = ws.cell(row=row, column=COL_CYCLE_TIME, value=cycle_time)
    tc_cell.alignment = Alignment(horizontal='center', vertical='top', wrap_text=True)
    tc_cell.font = Font(name='Calibri', size=11)

    # 7. Points clés (AH:AP)
    kp_cell = ws.cell(row=row, column=COL_KEY_POINTS_START, value=key_points)
    kp_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    kp_cell.font = Font(name='Calibri', size=10)

    # 8. Raison du point clé (AQ:AY)
    reason = data.get("raison_point_cle", "")
    reason_cell = ws.cell(row=row, column=COL_REASON_START, value=reason)
    reason_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    reason_cell.font = Font(name='Calibri', size=10)

    # Calculate auto row height based on longest text
    texts = [description, main_step, key_points, reason]
    max_text_len = max(len(str(t)) if t else 0 for t in texts)
    auto_height = calculate_row_height(max_text_len)
    ws.row_dimensions[row].height = auto_height


def copy_row_with_formatting(source_ws, target_ws, source_row, target_row):
    """Copie une ligne avec son formatage."""
    for col in range(1, 52):
        source_cell = source_ws.cell(row=source_row, column=col)
        target_cell = target_ws.cell(row=target_row, column=col)

        target_cell.value = source_cell.value

        if source_cell.has_style:
            target_cell.font = source_cell.font.copy()
            target_cell.border = source_cell.border.copy()
            target_cell.fill = source_cell.fill.copy()
            target_cell.number_format = source_cell.number_format
            target_cell.protection = source_cell.protection.copy()
            target_cell.alignment = source_cell.alignment.copy()

    for merged_range in source_ws.merged_cells.ranges:
        if merged_range.min_row == source_row and merged_range.max_row == source_row:
            new_range = f"{merged_range.min_col_letter}{target_row}:{merged_range.max_col_letter}{target_row}"
            target_ws.merge_cells(new_range)


def add_new_page(wb, page_num: int):
    """Ajoute une nouvelle page avec le meme header que la page 1."""
    source_ws = wb[TEMPLATE_SHEET]
    new_sheet_name = f"Page_{page_num}"

    new_ws = wb.copy_worksheet(source_ws)
    new_ws.title = new_sheet_name

    for row in range(FIRST_DATA_ROW, FIRST_DATA_ROW + MAX_ROWS_PER_PAGE + 10):
        for col in range(1, 52):
            cell = new_ws.cell(row=row, column=col)
            cell.value = None
            cell.font = Font(name='Calibri', size=11)
            cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

    for col in range(1, 52):
        header_cell = source_ws.cell(row=HEADER_ROW, column=col)
        new_header_cell = new_ws.cell(row=HEADER_ROW, column=col)
        new_header_cell.value = header_cell.value
        if header_cell.has_style:
            new_header_cell.font = header_cell.font.copy()
            new_header_cell.alignment = header_cell.alignment.copy()

    print(f"      Nouvelle page creee: {new_sheet_name}")
    return new_ws


_STOPWORDS = {"le", "la", "les", "un", "une", "des", "du", "de", "d", "sur", "dans",
              "avec", "en", "et", "ou", "a", "à", "pour", "par"}


def apply_keyword_formatting(cell, description, points_cles):
    """
    Met en ROUGE et GRAS, dans la description, les mots/segments qui sont
    repris dans le champ "Points cles" (points_cles), pour que la duplication
    entre les deux champs saute aux yeux a la lecture.

    Ex: description = "Prendre la piece avec la main droite et la deposer"
        points_cles  = "main droite"
        -> "main droite" apparait en rouge/gras dans la description.
    """
    if not description:
        cell.value = ""
        return

    from openpyxl.cell.text import InlineFont
    from openpyxl.cell.rich_text import TextBlock, CellRichText

    normal_font = InlineFont(rFont='Calibri', sz=10, b=False)
    red_bold_font = InlineFont(rFont='Calibri', sz=10, b=True, color=Color(rgb="FF0000"))

    desc = str(description)
    pts = str(points_cles or "")

    # 1. Extraction des termes de recherche a partir de points_cles
    raw_segments = [s.strip() for s in re.split(r'[/,]', pts) if s.strip()]
    search_terms = []

    for seg in raw_segments:
        if seg.lower() in desc.lower():
            search_terms.append(seg)
        else:
            for w in seg.split():
                clean_w = re.sub(r'[^\w]', '', w)
                if len(clean_w) >= 3 and clean_w.lower() not in _STOPWORDS:
                    search_terms.append(clean_w)

    if not search_terms:
        cell.value = desc
        return

    # 2. Localisation des intervalles correspondants dans la description
    intervals = []
    for term in search_terms:
        pattern = r'\b' + re.escape(term) + r'\b'
        matches = list(re.finditer(pattern, desc, flags=re.IGNORECASE))
        if not matches:
            matches = list(re.finditer(re.escape(term), desc, flags=re.IGNORECASE))
        for m in matches:
            intervals.append((m.start(), m.end()))

    if not intervals:
        cell.value = desc
        return

    # 3. Fusion des intervalles qui se chevauchent
    intervals.sort(key=lambda x: x[0])
    merged = []
    for s, e in intervals:
        if not merged:
            merged.append([s, e])
        else:
            if s <= merged[-1][1]:
                merged[-1][1] = max(merged[-1][1], e)
            else:
                merged.append([s, e])

    # 4. Construction du texte enrichi (rich text) openpyxl
    rich_blocks = []
    curr = 0
    for s, e in merged:
        if s > curr:
            rich_blocks.append(TextBlock(normal_font, desc[curr:s]))
        rich_blocks.append(TextBlock(red_bold_font, desc[s:e]))
        curr = e
    if curr < len(desc):
        rich_blocks.append(TextBlock(normal_font, desc[curr:]))

    cell.value = CellRichText(rich_blocks)


def convert_to_title_format(text):
    """Convertit un texte imperatif en format titre (nom d'action)."""
    if not text:
        return ""

    verb_mapping = {
        "prend": "Prise",
        "prends": "Prise",
        "coupe": "Coupe",
        "couper": "Coupe",
        "tranche": "Tranchage",
        "pose": "Pose",
        "poser": "Pose",
        "depose": "Depose",
        "plie": "Pliage",
        "plier": "Pliage",
        "replie": "Repliage",
        "tire": "Tirage",
        "tirer": "Tirage",
        "hisse": "Hissage",
        "souleve": "Soulevement",
        "soulever": "Soulevement",
        "tourne": "Rotation",
        "tourner": "Rotation",
        "pivote": "Pivotement",
        "retourne": "Retournement",
        "attend": "Attente",
        "attends": "Attente",
        "verifie": "Verification",
        "verifier": "Verification",
        "controle": "Controle",
        "controler": "Controle",
        "inspecte": "Inspection",
        "inspecter": "Inspection",
        "met": "Mise en place",
        "mets": "Mise en place",
        "mettre": "Mise en place",
        "tiens": "Maintien",
        "tenir": "Maintien",
        "maintiens": "Maintien",
        "deplace": "Deplacement",
        "deplacer": "Deplacement",
        "transfere": "Transfert",
        "transferer": "Transfert",
        "lisse": "Lissage",
        "lisser": "Lissage",
        "aplatis": "Aplatissement",
        "aplatir": "Aplatissement",
        "aligne": "Alignement",
        "aligner": "Alignement",
        "ajuste": "Ajustement",
        "ajuster": "Ajustement",
        "guide": "Guidage",
        "guider": "Guidage",
        "dirige": "Direction",
        "diriger": "Direction",
        "oriente": "Orientation",
        "orienter": "Orientation",
        "pousse": "Poussée",
        "pousser": "Poussée",
        "appuie": "Appui",
        "appuyer": "Appui",
        "frotte": "Frottement",
        "frotter": "Frottement",
        "glisse": "Glissement",
        "glisser": "Glissement",
        "introduis": "Introduction",
        "introduire": "Introduction",
        "enfile": "Enfilage",
        "enfiler": "Enfilage",
        "agrippe": "Agrippage",
        "agripper": "Agrippage",
        "attrape": "Attrapage",
        "attraper": "Attrapage",
        "ramasse": "Ramassage",
        "ramasser": "Ramassage",
        "reprend": "Reprise",
        "reprendre": "Reprise",
        "assemble": "Assemblage",
        "assembler": "Assemblage",
        "conditionne": "Conditionnement",
        "conditionner": "Conditionnement",
        "stocke": "Stockage",
        "stocker": "Stockage",
    }

    text_lower = text.lower().strip()
    words = text_lower.split()

    if not words:
        return text

    first_word = words[0]
    for verb_key, verb_title in verb_mapping.items():
        if first_word.startswith(verb_key):
            rest = " ".join(words[1:])
            if rest:
                return f"{verb_title} {rest}"
            return verb_title

    return text.title()


def get_merged_cell_value(ws, cell_ref):
    """Retourne la cellule reelle (non fusionnee) pour une reference donnee."""
    cell = ws[cell_ref]
    if isinstance(cell, openpyxl.cell.cell.MergedCell):
        for merged_range in ws.merged_cells.ranges:
            if cell.coordinate in merged_range:
                return ws.cell(row=merged_range.min_row, column=merged_range.min_col)
    return cell


def set_cell_value_safe(ws, cell_ref, value):
    """Ecrit dans une cellule meme si elle est fusionnee."""
    cell = get_merged_cell_value(ws, cell_ref)
    cell.value = value
    return cell


def fill_header_info(ws, project_name: str, video_name: str, total_steps: int):
    """Remplit les infos d'en-tete du template."""
    try:
        set_cell_value_safe(ws, 'B12', f"Analyse video - {video_name}")
    except Exception:
        pass
    try:
        set_cell_value_safe(ws, 'B14', f"Projet: {project_name}")
    except Exception:
        pass
    try:
        set_cell_value_safe(ws, 'B16', f"Nombre d'etapes analysees: {total_steps}")
    except Exception:
        pass


def generate_sos_excel(json_path: str, output_path: str = None, project_name: str = None):
    """Genere le fichier Excel SOS rempli a partir du rapport JSON."""
    json_path = Path(json_path)
    if not json_path.exists():
        raise FileNotFoundError(f"Rapport JSON non trouve : {json_path}")

    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)

    if not data:
        raise ValueError("Le rapport JSON est vide.")

    if output_path is None:
        output_path = json_path.parent / f"{json_path.parent.name}_SOS_Analysis.xlsx"
    else:
        output_path = Path(output_path)

    print(f"[1/4] Copie du template vers {output_path}...")
    copy_template_to_output(output_path)

    wb = load_workbook(output_path, rich_text=True)
    ws = wb[TEMPLATE_SHEET]

    video_name = json_path.parent.name
    if project_name is None:
        project_name = video_name

    print(f"[2/4] Remplissage des infos d'en-tete...")
    fill_header_info(ws, project_name, video_name, len(data))

    print(f"[3/4] Remplissage des {len(data)} operations...")
    current_page = 1
    current_ws = ws
    current_row = FIRST_DATA_ROW

    for i, step_data in enumerate(data, 1):
        if current_row >= FIRST_DATA_ROW + MAX_ROWS_PER_PAGE:
            current_page += 1
            print(f"      Creation de la page {current_page}...")
            current_ws = add_new_page(wb, current_page)
            current_row = FIRST_DATA_ROW

        image_path = step_data.get("frame_image_path", None)
        if image_path:
            image_path = image_path.replace("\\", "/")
            img_path_obj = Path(image_path)

            possible_paths = [
                json_path.parent.parent / image_path,
                json_path.parent / img_path_obj.name,
                json_path.parent / img_path_obj,
                Path(image_path),
                json_path.parent.parent / img_path_obj.name,
            ]

            img_full_path = None
            for p in possible_paths:
                if p.exists():
                    img_full_path = p
                    break

            image_path = str(img_full_path) if img_full_path else None

        fill_operation_row(current_ws, current_row, i, step_data, image_path)
        print(f"      Etape {i}/{len(data)} -> Ligne {current_row} (Page {current_page})")

        current_row += 1

    print(f"[4/4] Sauvegarde du fichier Excel...")
    wb.save(output_path)
    print(f"✅ Fichier genere avec succes : {output_path}")
    print(f"   Total: {len(data)} etapes sur {current_page} page(s)")

    return output_path


if __name__ == "__main__":
    import sys
    if len(sys.argv) > 1:
        json_file = sys.argv[1]
        generate_sos_excel(json_file)
    else:
        print("Usage: python excel_generator.py <path/to/rapport_analyse.json>")
        print("Example: python excel_generator.py outputs/video_9/rapport_analyse.json")